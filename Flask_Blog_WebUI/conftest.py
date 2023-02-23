from selenium import webdriver
import pytest
import os
from openpyxl import load_workbook


DATA_PATH = join_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')


@pytest.fixture(scope='module', autouse=True)
def broswer():
    broswer = webdriver.Chrome()
    yield broswer
    print('结束')
    broswer.quit()


def read_data(filename):
    data = load_workbook(os.path.join(DATA_PATH, filename)).active
    yield from data.iter_rows(min_row=2, values_only=True)


if __name__ == '__main__':
    values = read_data('index_navbar_text.xlsx')
    for value in values:
        print(value)
